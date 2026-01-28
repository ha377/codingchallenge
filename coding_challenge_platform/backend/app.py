import os
import random
import requests
import difflib
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from pymongo import MongoClient
from openpyxl import load_workbook

# =====================================================
# APP INITIALIZATION (MUST BE FIRST)
# =====================================================
app = Flask(__name__, static_folder="../frontend", static_url_path="")
CORS(app)

# =====================================================
# CONFIGURATION
# =====================================================
UPLOAD_FOLDER = "uploads"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "adminpass"

MONGO_URI = os.environ.get("MONGO_URI")
mongo_client = MongoClient(MONGO_URI)
DB_NAME = "coding_challenge"
SCORES_COLLECTION_NAME = "scores"

# =====================================================
# DATABASE SETUP
# =====================================================
try:
    mongo_client = MongoClient(MONGO_URI)
    mongo_db = mongo_client[DB_NAME]
    scores_collection = mongo_db[SCORES_COLLECTION_NAME]
    users_collection = mongo_db["users"]
    print("Successfully connected to MongoDB.")
except Exception as e:
    print("MongoDB connection error:", e)
    scores_collection = None
    users_collection = None

# =====================================================
# INITIALIZE DIRECTORIES
# =====================================================
def init_db():
    folders = [
        "mcq",
        "scramble/py", "scramble/c", "scramble/cpp", "scramble/java",
        "debug/py", "debug/c", "debug/cpp", "debug/java",
        "scramble_submissions/py", "scramble_submissions/c", "scramble_submissions/cpp", "scramble_submissions/java",
        "debug_submissions/correct/py", "debug_submissions/correct/c", "debug_submissions/correct/cpp", "debug_submissions/correct/java",
        "debug_submissions/wrong/py", "debug_submissions/wrong/c", "debug_submissions/wrong/cpp", "debug_submissions/wrong/java",
        "frontend_submissions",
        "Round2", "Round3", "Round4"
    ]
    for folder in folders:
        os.makedirs(os.path.join(UPLOAD_FOLDER, folder), exist_ok=True)

# =====================================================
# JUDGE0 HELPER
# =====================================================
def judge0_compile(source_code, language_id, stdin=""):
    url = "https://ce.judge0.com/submissions/?base64_encoded=false&wait=true"
    payload = {
        "source_code": source_code,
        "language_id": language_id,
        "stdin": stdin
    }
    try:
        r = requests.post(url, json=payload, timeout=10)
        if r.status_code in [200, 201]:
            res = r.json()
            return {
                "stdout": res.get("stdout", ""),
                "stderr": res.get("stderr", ""),
                "status": res.get("status", {}).get("description", "")
            }
        return {"error": f"Judge0 error {r.status_code}"}
    except Exception as e:
        return {"error": str(e)}

# =====================================================
# FRONTEND
# =====================================================
@app.route("/")
def index():
    return send_from_directory("../frontend", "index.html")

@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory("../frontend", path)

# =====================================================
# AUTH
# =====================================================
@app.route("/admin_login", methods=["POST"])
def admin_login():
    data = request.get_json()
    if data.get("username") == ADMIN_USERNAME and data.get("password") == ADMIN_PASSWORD:
        return jsonify({"message": "Admin login successful"}), 200
    return jsonify({"message": "Invalid credentials"}), 401

@app.route("/student_signup", methods=["POST"])
def student_signup():
    data = request.get_json()
    if users_collection.find_one({"username": data["username"]}):
        return jsonify({"message": "User exists"}), 400
    users_collection.insert_one({
        "username": data["username"],
        "password": data["password"],
        "role": "student",
        "created_at": datetime.now()
    })
    return jsonify({"message": "Signup successful"}), 201

@app.route("/student_login", methods=["POST"])
def student_login():
    data = request.get_json()
    user = users_collection.find_one({"username": data["username"], "role": "student"})
    if user and user["password"] == data["password"]:
        return jsonify({"message": "Login successful"}), 200
    return jsonify({"message": "Invalid login"}), 401

# =====================================================
# ROUND 1 – MCQ
# =====================================================
@app.route("/get_mcq_questions")
def get_mcq_questions():
    file_path = os.path.join(UPLOAD_FOLDER, "mcq", "questions.xlsx")
    if not os.path.exists(file_path):
        return jsonify({"error": "questions.xlsx missing"}), 404

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        raw_headers = [cell.value for cell in ws[1]]
        
        # Normalize headers to expected format
        header_map = {}
        for i, header in enumerate(raw_headers):
            if header:
                h = str(header).strip().lower()
                if 'id' in h or 'question_id' in h or 'q_id' in h:
                    header_map[i] = 'id'
                elif 'question' in h and 'text' in h:
                    header_map[i] = 'question_text'
                elif 'question' in h:
                    header_map[i] = 'question_text'
                elif 'option' in h and 'a' in h:
                    header_map[i] = 'optionA'
                elif 'option' in h and 'b' in h:
                    header_map[i] = 'optionB'
                elif 'option' in h and 'c' in h:
                    header_map[i] = 'optionC'
                elif 'option' in h and 'd' in h:
                    header_map[i] = 'optionD'
                elif 'correct' in h or 'answer' in h:
                    header_map[i] = 'correct_answer'
                else:
                    header_map[i] = header
        
        questions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                question = {}
                for i, value in enumerate(row):
                    if i in header_map:
                        question[header_map[i]] = value
                
                # Ensure required fields exist
                if 'question_text' in question and question['question_text']:
                    if 'id' not in question:
                        question['id'] = len(questions) + 1
                    questions.append(question)
        
        return jsonify(questions)
    except Exception as e:
        return jsonify({"error": f"Error reading Excel file: {str(e)}"}), 500

# =====================================================
# ADMIN FILE UPLOAD
# =====================================================
@app.route("/admin_upload", methods=["POST"])
def admin_upload():
    try:
        round_type = request.form.get("round")
        lang = request.form.get("lang", "py")
        file = request.files.get("file")
        
        if not file:
            return jsonify({"message": "No file provided"}), 400
        
        # Handle MCQ upload
        if round_type == "mcq":
            if not file.filename.endswith(".xlsx"):
                return jsonify({"message": "MCQ file must be .xlsx"}), 400
            save_path = os.path.join(UPLOAD_FOLDER, "mcq", "questions.xlsx")
            file.save(save_path)
            return jsonify({"message": "MCQ file uploaded successfully"}), 200
        
        # Handle code rounds (scramble/debug)
        elif round_type in ["scramble", "debug"]:
            # Get file extension from uploaded file
            ext = os.path.splitext(file.filename)[1]
            if not ext:
                # Fallback to lang if no extension
                ext_map = {"py": ".py", "c": ".c", "cpp": ".cpp", "java": ".java"}
                ext = ext_map.get(lang, ".txt")
            
            # Create directory if needed
            folder = os.path.join(UPLOAD_FOLDER, round_type, lang)
            os.makedirs(folder, exist_ok=True)
            
            # Save with original filename (preserving extension)
            save_path = os.path.join(folder, file.filename)
            file.save(save_path)
            return jsonify({"message": f"{round_type.capitalize()} code uploaded successfully"}), 200
        
        return jsonify({"message": "Invalid round type"}), 400
        
    except Exception as e:
        return jsonify({"message": f"Upload error: {str(e)}"}), 500

# =====================================================
# GET CODE FILE LISTS
# =====================================================
@app.route("/get_scrambled_code_list")
def get_scrambled_code_list():
    lang = request.args.get("lang", "py")
    folder = os.path.join(UPLOAD_FOLDER, "scramble", lang)
    if not os.path.exists(folder):
        return jsonify([])
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
    return jsonify(files)

@app.route("/get_buggy_code_list")
def get_buggy_code_list():
    lang = request.args.get("lang", "py")
    folder = os.path.join(UPLOAD_FOLDER, "debug", lang)
    if not os.path.exists(folder):
        return jsonify([])
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
    return jsonify(files)

@app.route("/admin/code_questions")
def admin_code_questions():
    round_type = request.args.get("round")
    lang = request.args.get("lang", "py")
    folder = os.path.join(UPLOAD_FOLDER, round_type, lang)
    if not os.path.exists(folder):
        return jsonify([])
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
    return jsonify(files)

# =====================================================
# GET CODE FILE CONTENT
# =====================================================
@app.route("/get_scrambled_code")
def get_scrambled_code():
    lang = request.args.get("lang", "py")
    filename = request.args.get("file")
    file_path = os.path.join(UPLOAD_FOLDER, "scramble", lang, filename)
    
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    
    # Scramble the lines
    scrambled = lines.copy()
    random.shuffle(scrambled)
    
    return jsonify({
        "original": "".join(lines),
        "scrambled": "".join(scrambled)
    })

@app.route("/get_buggy_code")
def get_buggy_code():
    lang = request.args.get("lang", "py")
    filename = request.args.get("file")
    file_path = os.path.join(UPLOAD_FOLDER, "debug", lang, filename)
    
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    
    with open(file_path, "r", encoding="utf-8") as f:
        code = f.read()
    
    return jsonify({"code": code})

@app.route("/admin/file_content")
def admin_file_content():
    file_path = request.args.get("path")
    full_path = os.path.join(UPLOAD_FOLDER, file_path)
    
    if not os.path.exists(full_path):
        return jsonify({"error": "File not found"}), 404
    
    with open(full_path, "r", encoding="utf-8") as f:
        content = f.read()
    
    return jsonify({"content": content})

# =====================================================
# SUBMIT SCORES AND CODE
# =====================================================
@app.route("/submit_mcq_score", methods=["POST"])
def submit_mcq_score():
    try:
        data = request.get_json()
        scores_collection.update_one(
            {"username": data["username"], "round": "MCQ"},
            {"$set": {
                "score": data["score"],
                "percentage": data.get("percentage", 0),
                "total_questions": data.get("total_questions", 10),
                "timestamp": datetime.now()
            }},
            upsert=True
        )
        return jsonify({"message": "Score saved"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/submit_scrambled_code", methods=["POST"])
def submit_scrambled_code():
    try:
        data = request.get_json()
        student_code = data.get("code", "")
        lang = data.get("lang", "py")
        username = data.get("username")
        file_path = data.get("file_path")
        remaining_time = data.get("remaining_time", 0)  # Time remaining when submitted
        
        # Read original code
        original_path = os.path.join(UPLOAD_FOLDER, file_path)
        if not os.path.exists(original_path):
            return jsonify({"error": "Original file not found"}), 404
        
        with open(original_path, "r", encoding="utf-8") as f:
            original_code = f.read()
        
        # Calculate similarity
        similarity = difflib.SequenceMatcher(None, original_code, student_code).ratio()
        score = int(similarity * 100)
        
        # Create folders for scramble submissions
        save_folder = os.path.join(UPLOAD_FOLDER, "scramble_submissions", lang)
        os.makedirs(save_folder, exist_ok=True)
        
        # Get file extension
        ext_map = {"py": ".py", "c": ".c", "cpp": ".cpp", "java": ".java"}
        ext = ext_map.get(lang, ".txt")
        
        # Save student code with username prefix
        filename = f"{username}_{os.path.basename(file_path).replace(ext, '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        save_path = os.path.join(save_folder, filename)
        
        with open(save_path, "w", encoding="utf-8") as f:
            f.write(student_code)
        
        # Save score
        scores_collection.update_one(
            {"username": username, "round": "Scramble"},
            {"$set": {
                "score": score,
                "language": lang,
                "file": file_path,
                "saved_file": filename,
                "similarity": f"{similarity:.2%}",
                "remaining_time": remaining_time,
                "timestamp": datetime.now()
            }},
            upsert=True
        )
        
        return jsonify({
            "message": "Code submitted successfully",
            "filename": filename
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/submit_debug_code", methods=["POST"])
def submit_debug_code():
    try:
        data = request.get_json()
        student_code = data.get("code", "")
        lang = data.get("lang", "py")
        username = data.get("username")
        file_path = data.get("file_path")
        remaining_time = data.get("remaining_time", 0)  # Time remaining when submitted
        
        # Test the code with Judge0 to see if it compiles and runs correctly
        lang_map = {"py": 71, "c": 50, "cpp": 54, "java": 62}
        compile_result = judge0_compile(student_code, lang_map.get(lang, 71), "")
        
        # Determine if code is correct based on compilation
        is_correct = False
        if compile_result.get("status") == "Accepted" and not compile_result.get("stderr") and not compile_result.get("error"):
            is_correct = True
        
        # Create folders for correct/wrong submissions
        status_folder = "correct" if is_correct else "wrong"
        save_folder = os.path.join(UPLOAD_FOLDER, "debug_submissions", status_folder, lang)
        os.makedirs(save_folder, exist_ok=True)
        
        # Get file extension
        ext_map = {"py": ".py", "c": ".c", "cpp": ".cpp", "java": ".java"}
        ext = ext_map.get(lang, ".txt")
        
        # Save student code with username prefix
        filename = f"{username}_{os.path.basename(file_path).replace(ext, '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        save_path = os.path.join(save_folder, filename)
        
        with open(save_path, "w", encoding="utf-8") as f:
            f.write(student_code)
        
        # Save score - 100 if correct, 0 if wrong
        score = 100 if is_correct else 0
        scores_collection.update_one(
            {"username": username, "round": "Debugging"},
            {"$set": {
                "score": score,
                "language": lang,
                "file": file_path,
                "status": status_folder,
                "saved_file": filename,
                "remaining_time": remaining_time,
                "timestamp": datetime.now()
            }},
            upsert=True
        )
        
        return jsonify({
            "message": "Code submitted successfully",
            "status": status_folder,
            "filename": filename
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/submit_frontend", methods=["POST"])
def submit_frontend():
    try:
        file = request.files.get("file")
        username = request.form.get("username")
        
        if not file or not username:
            return jsonify({"message": "Missing file or username"}), 400
        
        # Create user folder
        user_folder = os.path.join(UPLOAD_FOLDER, "frontend_submissions", username)
        os.makedirs(user_folder, exist_ok=True)
        
        # Save file
        save_path = os.path.join(user_folder, file.filename)
        file.save(save_path)
        
        return jsonify({"message": "File uploaded successfully"}), 200
        
    except Exception as e:
        return jsonify({"message": f"Upload error: {str(e)}"}), 500

# =====================================================
# ROUND 3 – CHECK DEBUG CODE (JUDGE0)
# =====================================================
@app.route("/check_debug_code", methods=["POST"])
def check_debug_code():
    data = request.get_json()
    lang_map = {"py": 71, "c": 50, "cpp": 54, "java": 62}
    result = judge0_compile(
        data.get("code"),
        lang_map.get(data.get("lang", "py"), 71),
        data.get("input", "")
    )
    return jsonify(result)

# =====================================================
# STUDENT SCORES
# =====================================================
@app.route("/student/scores")
def get_student_scores():
    username = request.args.get("username")
    if not username:
        return jsonify({"error": "Username required"}), 400
    
    scores = list(scores_collection.find({"username": username}, {"_id": 0}))
    return jsonify(scores)

# =====================================================
# ADMIN ENDPOINTS
# =====================================================
@app.route("/admin/scores")
def get_all_scores():
    scores = list(scores_collection.find({}, {"_id": 0}))
    return jsonify(scores)

@app.route("/admin/questions")
def get_admin_questions():
    file_path = os.path.join(UPLOAD_FOLDER, "mcq", "questions.xlsx")
    if not os.path.exists(file_path):
        return jsonify([])
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        raw_headers = [cell.value for cell in ws[1]]
        
        # Normalize headers
        header_map = {}
        for i, header in enumerate(raw_headers):
            if header:
                h = str(header).strip().lower()
                if 'id' in h or 'question_id' in h or 'q_id' in h:
                    header_map[i] = 'id'
                elif 'question' in h and 'text' in h:
                    header_map[i] = 'question_text'
                elif 'question' in h:
                    header_map[i] = 'question_text'
                elif 'option' in h and 'a' in h:
                    header_map[i] = 'optionA'
                elif 'option' in h and 'b' in h:
                    header_map[i] = 'optionB'
                elif 'option' in h and 'c' in h:
                    header_map[i] = 'optionC'
                elif 'option' in h and 'd' in h:
                    header_map[i] = 'optionD'
                elif 'correct' in h or 'answer' in h:
                    header_map[i] = 'correct_answer'
                else:
                    header_map[i] = header
        
        questions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                question = {}
                for i, value in enumerate(row):
                    if i in header_map:
                        question[header_map[i]] = value
                
                if 'question_text' in question and question['question_text']:
                    if 'id' not in question:
                        question['id'] = len(questions) + 1
                    questions.append(question)
        
        return jsonify(questions)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/submissions")
def get_admin_submissions():
    submissions = []
    
    # Get all scores from MongoDB for metadata lookup
    all_scores = list(scores_collection.find({}))
    
    # Frontend submissions
    base_folder = os.path.join(UPLOAD_FOLDER, "frontend_submissions")
    if os.path.exists(base_folder):
        for username in os.listdir(base_folder):
            user_folder = os.path.join(base_folder, username)
            if os.path.isdir(user_folder):
                for filename in os.listdir(user_folder):
                    submissions.append({
                        "team": username,
                        "round": "Frontend Challenge",
                        "filename": filename,
                        "path": f"frontend_submissions/{username}/{filename}"
                    })
    
    # Scramble submissions
    scramble_folder = os.path.join(UPLOAD_FOLDER, "scramble_submissions")
    if os.path.exists(scramble_folder):
        for lang in os.listdir(scramble_folder):
            lang_folder = os.path.join(scramble_folder, lang)
            if os.path.isdir(lang_folder):
                for filename in os.listdir(lang_folder):
                    username = filename.split('_')[0]
                    file_path = os.path.join(lang_folder, filename)
                    
                    # Look up remaining_time and timestamp from MongoDB
                    score_data = next((s for s in all_scores if s.get('username') == username and s.get('round') == 'Scramble' and s.get('saved_file') == filename), None)
                    remaining_time = score_data.get('remaining_time') if score_data else None
                    timestamp = score_data.get('timestamp').strftime('%Y-%m-%d %H:%M:%S') if score_data and score_data.get('timestamp') else 'N/A'
                    file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                    
                    submissions.append({
                        "team": username,
                        "round": "Code Scramble",
                        "filename": filename,
                        "language": lang,
                        "remaining_time": remaining_time,
                        "timestamp": timestamp,
                        "size": file_size,
                        "path": f"scramble_submissions/{lang}/{filename}"
                    })
    
    # Debug submissions - correct
    debug_correct_folder = os.path.join(UPLOAD_FOLDER, "debug_submissions", "correct")
    if os.path.exists(debug_correct_folder):
        for lang in os.listdir(debug_correct_folder):
            lang_folder = os.path.join(debug_correct_folder, lang)
            if os.path.isdir(lang_folder):
                for filename in os.listdir(lang_folder):
                    # Extract username from filename (format: username_problem_timestamp.ext)
                    username = filename.split('_')[0]
                    file_path = os.path.join(lang_folder, filename)
                    
                    # Look up remaining_time and timestamp from MongoDB
                    score_data = next((s for s in all_scores if s.get('username') == username and s.get('round') == 'Debugging' and s.get('saved_file') == filename), None)
                    remaining_time = score_data.get('remaining_time') if score_data else None
                    timestamp = score_data.get('timestamp').strftime('%Y-%m-%d %H:%M:%S') if score_data and score_data.get('timestamp') else 'N/A'
                    file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                    
                    submissions.append({
                        "team": username,
                        "round": "Debug - Correct",
                        "filename": filename,
                        "language": lang,
                        "remaining_time": remaining_time,
                        "timestamp": timestamp,
                        "size": file_size,
                        "path": f"debug_submissions/correct/{lang}/{filename}"
                    })
    
    # Debug submissions - wrong
    debug_wrong_folder = os.path.join(UPLOAD_FOLDER, "debug_submissions", "wrong")
    if os.path.exists(debug_wrong_folder):
        for lang in os.listdir(debug_wrong_folder):
            lang_folder = os.path.join(debug_wrong_folder, lang)
            if os.path.isdir(lang_folder):
                for filename in os.listdir(lang_folder):
                    username = filename.split('_')[0]
                    file_path = os.path.join(lang_folder, filename)
                    
                    # Look up remaining_time and timestamp from MongoDB
                    score_data = next((s for s in all_scores if s.get('username') == username and s.get('round') == 'Debugging' and s.get('saved_file') == filename), None)
                    remaining_time = score_data.get('remaining_time') if score_data else None
                    timestamp = score_data.get('timestamp').strftime('%Y-%m-%d %H:%M:%S') if score_data and score_data.get('timestamp') else 'N/A'
                    file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                    
                    submissions.append({
                        "team": username,
                        "round": "Debug - Wrong",
                        "filename": filename,
                        "language": lang,
                        "remaining_time": remaining_time,
                        "timestamp": timestamp,
                        "size": file_size,
                        "path": f"debug_submissions/wrong/{lang}/{filename}"
                    })
    
    return jsonify(submissions)

@app.route("/admin/submission_file")
def get_submission_file():
    file_path = request.args.get("path")
    full_path = os.path.join(UPLOAD_FOLDER, file_path)
    
    if not os.path.exists(full_path):
        return jsonify({"error": "File not found"}), 404
    
    with open(full_path, "r", encoding="utf-8") as f:
        content = f.read()
    
    return jsonify({"content": content})

@app.route("/admin/scores/delete", methods=["DELETE"])
def delete_all_scores():
    try:
        scores_collection.delete_many({})
        return jsonify({"message": "All scores deleted"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =====================================================
# SERVER START
# =====================================================
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 8000))
    app.run(debug=True, host="0.0.0.0", port=port)

